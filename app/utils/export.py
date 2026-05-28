"""Export leads to CSV or Excel."""
import csv, io
from typing import List
from app.models.schemas import BusinessLead

FIELDS = [
    "name","category","rating","reviews_count","address","city","state",
    "postal_code","phone","email","website","maps_url","latitude","longitude",
    "is_open_now","description","place_id","data_quality_score","scraped_at",
]

def leads_to_csv(leads: List[BusinessLead]) -> bytes:
    buf = io.StringIO()
    writer = csv.DictWriter(buf, fieldnames=FIELDS, extrasaction="ignore")
    writer.writeheader()
    for lead in leads:
        row = lead.model_dump(mode="json")
        # flatten lists
        for f in ("opening_hours","images","services","amenities"):
            v = row.get(f)
            row[f] = " | ".join(v) if isinstance(v,list) else ""
        row["social_links"] = str(row.get("social_links","") or "")
        writer.writerow(row)
    return buf.getvalue().encode("utf-8-sig")  # BOM for Excel

def leads_to_excel(leads: List[BusinessLead]) -> bytes:
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
    except ImportError:
        # Fallback to CSV if openpyxl not installed
        return leads_to_csv(leads)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Leads"

    # Header style
    hdr_fill = PatternFill("solid", fgColor="1E40AF")
    hdr_font = Font(bold=True, color="FFFFFF", size=11)

    ws.append(FIELDS)
    for col_idx, _ in enumerate(FIELDS, 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = hdr_fill
        cell.font = hdr_font
        cell.alignment = Alignment(horizontal="center")
        ws.column_dimensions[get_column_letter(col_idx)].width = 20

    for lead in leads:
        row = lead.model_dump(mode="json")
        for f in ("opening_hours","images","services","amenities"):
            v = row.get(f); row[f] = " | ".join(v) if isinstance(v,list) else ""
        row["social_links"] = str(row.get("social_links","") or "")
        ws.append([row.get(f,"") for f in FIELDS])

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
